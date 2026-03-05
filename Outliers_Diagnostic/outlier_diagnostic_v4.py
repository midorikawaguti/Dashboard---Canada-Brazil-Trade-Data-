"""
OUTLIER GROUPING DIAGNOSTIC v4 — outlier_grouping_diagnostic.xlsx
==================================================================
METHOD: Log-scale IQR + 4 classification guards

  Guard 1  Min log-std floor  (MIN_LOG_STD = 0.10)
           When all prices in a group are nearly identical, log-std → 0
           and tiny rounding differences produce enormous z-scores.
           Flooring log-std at 0.10 (≈ ±25% natural price band) prevents
           this. E.g. ethanolamine at CV=0.28% no longer fires outliers.

  Guard 2  CV suppression  (CV_SUPPRESS = 0.02)
           If the coefficient of variation of prices in a group is below
           2%, all prices are so homogeneous there are no meaningful
           outliers — all flags are suppressed for that group.

  Guard 3  Outlier rate cap  (MAX_OUTLIER_PCT = 0.15)
           If more than 15% of a group's rows are flagged, the
           distribution is likely bimodal / structurally heterogeneous
           (e.g. HS 3507.90.00 covers both $1/kg bulk and $56k/kg
           specialty enzymes). Flags are demoted to REVIEW instead of
           MEDIUM/HIGH/CRITICAL.

  Guard 4  Percentile-rank severity
           Replaces hard logz ≥ 2/3/5 thresholds, which create false
           precision at tier boundaries (e.g. $3,355 = CRITICAL at
           z=3.006 vs $3,298 = HIGH at z=2.996 despite 1.7% price diff).
           Severity is now relative to other flagged rows in the same
           group: top 10% by |logz| = CRITICAL, next 20% = HIGH, 70% = MEDIUM.
SPEED OPTIMISATIONS vs v1:
  - Data rows written with ws.append() — no per-cell font/fill loops
  - Colours applied as conditional formatting rules (one call per column)
  - Detail sheets use pandas ExcelWriter bulk write, then header overlay
  - No per-row Python loops touching openpyxl cells

Expected runtime on 2.5M rows: 3-6 min total (vs 20+ min in v1)

INPUT  : Dataset/Normalised/prices_normalised.csv
OUTPUT : Dataset/Reports/outlier_grouping_diagnostic.xlsx

Run:
  python outlier_diagnostic.py
  python outlier_diagnostic.py --input /path/to/prices_normalised.csv
"""

import argparse, time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              GradientFill)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── CONFIG ────────────────────────────────────────────────────────────────────
BASE_DIR    = Path.cwd()
INPUT_FILE  = BASE_DIR / "Dataset" / "Normalised" / "prices_normalised.csv"
OUTPUT_DIR  = BASE_DIR / "Dataset" / "Reports"
OUTPUT_FILE = OUTPUT_DIR / "outlier_grouping_diagnostic_v4.xlsx"

# Log-IQR multiplier — applied to IQR of log10(unit_price).
# 1.5× keeps both fences positive for typical trade-price distributions.
LOG_IQR_MULT = 1.5
MIN_ROWS = 5

# ── Outlier classification guards (v4) ───────────────────────────────────────
# Guard 1 — min log-std floor: prevents z-score explosion on near-uniform groups
MIN_LOG_STD = 0.10       # 0.10 log-units ≈ ±25% natural price band

# Guard 2 — CV suppression: fully suppress flags when prices are homogeneous
CV_SUPPRESS = 0.02       # 2% CV threshold

# Guard 3 — outlier rate cap: bimodal/heterogeneous groups get REVIEW tier
MAX_OUTLIER_PCT = 0.15   # 15% cap before downgrading to REVIEW

# Guard 4 — percentile-rank severity (within each group's outlier population)
SEV_CRITICAL_PCT = 0.10  # top 10% by |logz| → CRITICAL
SEV_HIGH_PCT     = 0.20  # next 20%           → HIGH  (rest → MEDIUM)

GROUPINGS = {
    "G1_Commodity":              ["Commodity"],
    "G2_Comm_Country":           ["Commodity", "Country"],
    "G3_Comm_Province":          ["Commodity", "Province"],
    "G4_Comm_Country_Province":  ["Commodity", "Country", "Province"],
    "G5_Comm_Season":            ["Commodity", "_season"],
    "G6_Comm_Country_Prov_Seas": ["Commodity", "Country", "Province", "_season"],
}

# ── COLOURS ───────────────────────────────────────────────────────────────────
C_DARK="1F3864"; C_MID="2E75B6"; C_LIGHT="D6E4F0"; C_WHITE="FFFFFF"
C_AMBER="FFF2CC"; C_RED="FCE4D6"; C_GREEN="E2EFDA"; C_GREY="F2F2F2"
C_STRIPE="EBF3FB"; C_ORANGE="FFF0D9"
C_RED_D="C00000"; C_GREEN_D="375623"; C_AMBER_D="FFD966"

def xf(c):          return PatternFill("solid", fgColor=c)
def bf(c="000000", s=10): return Font(bold=True,  color=c, name="Arial", size=s)
def nf(c="000000", s=10): return Font(bold=False, color=c, name="Arial", size=s)
def xb():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)
def ac(): return Alignment(horizontal="center", vertical="center", wrap_text=False)
def ar(): return Alignment(horizontal="right",  vertical="center")
def al(i=1): return Alignment(horizontal="left", vertical="center", indent=i)
def awrap(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── STEP 1: LOAD ──────────────────────────────────────────────────────────────
def load(path: Path):
    print(f"  Loading {path.name}...", end=" ", flush=True)
    t0 = time.time()
    df = pd.read_csv(path, low_memory=False)
    valid = (
        df["_conversion_status"].isin(["OK", "CONVERTED", "ASSUMED"]) &
        df["_unit_price"].notna() &
        (df["_unit_price"] > 0) &
        (df["Value ($)"].fillna(0) > 0)
    )
    df_valid = df[valid].copy().reset_index(drop=True)
    df_valid["_row_id"] = df_valid.index  # stable integer ID for joining later
    print(f"{len(df):,} rows → {len(df_valid):,} valid  ({time.time()-t0:.1f}s)")
    return df, df_valid

# ── STEP 2: VECTORISED STATS ──────────────────────────────────────────────────
def compute_group_stats(df_valid: pd.DataFrame, group_cols: list) -> pd.DataFrame:
    t0 = time.time()
    g = df_valid.groupby(group_cols, dropna=False)

    agg = g["_unit_price"].agg(
        n_rows        = "count",
        mean_price    = "mean",
        median_price  = "median",
        std_price     = "std",
        min_price     = "min",
        max_price     = "max",
        q1            = lambda x: x.quantile(0.25),
        q3            = lambda x: x.quantile(0.75),
    ).reset_index()

    agg["iqr"]          = (agg["q3"] - agg["q1"]).round(2)
    agg["cv"]           = (agg["std_price"] / agg["median_price"].replace(0, np.nan)).round(2)
    agg["sufficient"]   = agg["n_rows"] >= MIN_ROWS

    # ── Log-IQR fences (v3) ──────────────────────────────────────────
    # Compute IQR on log10(unit_price), then convert fences back to
    # price-space. Lower fence is always positive; both fences are
    # symmetric in multiplicative terms.
    _lp = np.log10(df_valid["_unit_price"].clip(lower=1e-9))
    _lp_df = df_valid[group_cols].copy()
    _lp_df["_log_price"] = _lp.values
    log_agg = _lp_df.groupby(group_cols, dropna=False)["_log_price"].agg(
        log_q1  = lambda x: x.quantile(0.25),
        log_q3  = lambda x: x.quantile(0.75),
        log_med = "median",
        log_std = "std",
    ).reset_index()
    del _lp_df
    log_agg["log_iqr"]      = log_agg["log_q3"] - log_agg["log_q1"]
    log_agg["lower_fence"]  = (10 ** (log_agg["log_q1"] - LOG_IQR_MULT * log_agg["log_iqr"])).round(2)
    log_agg["upper_fence"]  = (10 ** (log_agg["log_q3"] + LOG_IQR_MULT * log_agg["log_iqr"])).round(2)
    log_agg["log_med_price"]= (10 ** log_agg["log_med"]).round(2)
    log_agg["log_std"]      = log_agg["log_std"].round(6)
    agg = agg.merge(
        log_agg[group_cols + ["lower_fence","upper_fence","log_iqr","log_med_price","log_std"]],
        on=group_cols, how="left"
    )

    for col in ["mean_price","median_price","std_price","min_price","max_price","q1","q3"]:
        agg[col] = agg[col].round(2)

    # Value and Quantity totals for context (stats above are all on unit price)
    val_qty = df_valid.groupby(group_cols, dropna=False).agg(
        total_value = ("Value ($)",       "sum"),
        total_qty   = ("_qty_normalised", "sum"),
    ).reset_index()
    val_qty["total_value"] = val_qty["total_value"].round(0).astype("Int64")
    val_qty["total_qty"]   = val_qty["total_qty"].round(2)
    agg = agg.merge(val_qty, on=group_cols, how="left")

    # Merge log-IQR fences back to rows — vectorised outlier detection
    merged = df_valid[group_cols + ["_unit_price", "_row_id"]].merge(
        agg[group_cols + ["lower_fence","upper_fence","sufficient",
                          "log_med_price","log_std"]],
        on=group_cols, how="left"
    )
    merged["_out"]   = ((merged["_unit_price"] < merged["lower_fence"]) |
                        (merged["_unit_price"] > merged["upper_fence"]))
    # z-score in log space — Guard 1: clip log_std to MIN_LOG_STD floor so that
    # near-zero variance (homogeneous prices) cannot inflate z-scores arbitrarily
    _log_p   = np.log10(merged["_unit_price"].clip(lower=1e-9))
    _log_med = np.log10(merged["log_med_price"].clip(lower=1e-9))
    _log_std = merged["log_std"].clip(lower=MIN_LOG_STD)
    merged["_z"] = ((_log_p - _log_med) / _log_std).abs()
    # Guard 2: suppress flags on groups where CV < threshold (homogeneous prices)
    cv_lookup = agg[group_cols + ["cv"]].copy()
    merged = merged.merge(cv_lookup, on=group_cols, how="left")
    # Only count outliers for sufficient groups
    merged.loc[~merged["sufficient"], "_out"] = False
    # Suppress: prices too uniform to have meaningful outliers
    merged.loc[merged["cv"].fillna(0) < CV_SUPPRESS, "_out"] = False
    # Legacy tier columns — kept only to populate the per-group count columns
    # in the detail sheets. Severity is reassigned via Guard 4 further below.
    merged["_out_m"] = merged["_out"] & (merged["_z"] >= 2)
    merged["_out_h"] = merged["_out"] & (merged["_z"] >= 3)
    merged["_out_c"] = merged["_out"] & (merged["_z"] >= 5)

    out_agg = merged.groupby(group_cols, dropna=False).agg(
        n_outlier_any  = ("_out",   "sum"),
        n_outlier_med  = ("_out_m", "sum"),
        n_outlier_hi   = ("_out_h", "sum"),
        n_outlier_crit = ("_out_c", "sum"),
    ).reset_index()

    result = agg.merge(out_agg, on=group_cols, how="left")
    result["pct_outlier"] = (result["n_outlier_any"] / result["n_rows"] * 100).round(2)
    result["pct_no_outl"] = (100 - result["pct_outlier"]).clip(0).round(2)
    result["too_few"]     = ~result["sufficient"]
    # Guard 3: mark groups where flagged rate exceeds cap → REVIEW tier
    result["heterogeneous"] = (
        result["sufficient"] &
        (result["pct_outlier"] / 100 > MAX_OUTLIER_PCT)
    )

    # Guards 3 + 4: assign severity to flagged rows
    het_lookup = result[group_cols + ["heterogeneous"]].copy()
    merged = merged.merge(het_lookup, on=group_cols, how="left")

    merged["_severity"] = ""
    flagged_mask = merged["_out"]

    if flagged_mask.any():
        # Guard 3: rows in heterogeneous groups → REVIEW
        het_mask    = flagged_mask & merged["heterogeneous"].fillna(False)
        merged.loc[het_mask, "_severity"] = "REVIEW"

        # Guard 4: percentile-rank severity within each group for all other flagged rows
        normal_mask = flagged_mask & ~het_mask
        if normal_mask.any():
            out_rows = merged.loc[normal_mask, group_cols + ["_z"]].copy()
            # rank descending (highest |z| = rank 1) as a percentile within the group
            out_rows["_rank_pct"] = out_rows.groupby(
                group_cols, dropna=False
            )["_z"].rank(method="max", ascending=False, pct=True)
            sev = np.where(
                out_rows["_rank_pct"] <= SEV_CRITICAL_PCT, "CRITICAL",
                np.where(
                    out_rows["_rank_pct"] <= SEV_CRITICAL_PCT + SEV_HIGH_PCT,
                    "HIGH", "MEDIUM"
                )
            )
            merged.loc[normal_mask, "_severity"] = sev
    merged["_z"] = merged["_z"].round(2)
    slim_cols = ["_row_id", "_severity", "_z", "log_med_price", "lower_fence", "upper_fence"]
    slim_cols = [col for col in slim_cols if col in merged.columns]
    flagged_rows = merged.loc[merged["_out"], slim_cols].copy()
    # free the large merged df immediately
    del merged

    t1 = time.time()
    return result, flagged_rows, t1 - t0

# ── STEP 3: BUILD OVERVIEW ────────────────────────────────────────────────────
def build_overview(all_stats: dict) -> pd.DataFrame:
    g1 = all_stats["G1_Commodity"][
        ["Commodity","n_rows","n_outlier_any","pct_outlier","too_few"]
    ].copy()
    g1.columns = ["Commodity","G1_valid_rows","G1_outliers","G1_pct_outlier","G1_too_few"]

    overview = g1.copy()

    label_map = {
        "G2_Comm_Country":           "G2",
        "G3_Comm_Province":          "G3",
        "G4_Comm_Country_Province":  "G4",
        "G5_Comm_Season":            "G5",
        "G6_Comm_Country_Prov_Seas": "G6",
    }

    for gname, df_g in all_stats.items():
        if gname == "G1_Commodity": continue
        p = label_map[gname]
        ca = df_g.groupby("Commodity", dropna=False).agg(
            n_grps    = ("Commodity",      "count"),
            n_too_few = ("too_few",        "sum"),
            n_out     = ("n_outlier_any",  "sum"),
            tot_rows  = ("n_rows",         "sum"),
        ).reset_index()
        ca[f"{p}_groups"]      = ca["n_grps"]
        ca[f"{p}_pct_too_few"] = (ca["n_too_few"] / ca["n_grps"]  * 100).round(1)
        ca[f"{p}_outliers"]    = ca["n_out"]
        ca[f"{p}_pct_outlier"] = (ca["n_out"] / ca["tot_rows"].replace(0,np.nan) * 100).round(2)
        keep = ["Commodity",
                f"{p}_groups", f"{p}_pct_too_few",
                f"{p}_outliers", f"{p}_pct_outlier"]
        overview = overview.merge(ca[keep], on="Commodity", how="left")

    # WHY column
    if "G2_Comm_Country" in all_stats:
        overview["WHY_G1_vs_G2"] = np.where(
            overview["G1_too_few"],
            "G1: TOO FEW — no baseline",
            np.where(
                overview["G2_pct_too_few"].fillna(0) > 50,
                "Fragmentation: >50% G2 groups TOO FEW",
                np.where(
                    overview["G2_pct_outlier"].fillna(overview["G1_pct_outlier"]) <
                    overview["G1_pct_outlier"] - 2,
                    "Better precision: outliers reduced in G2",
                    "Similar: grouping by country makes little difference"
                )
            )
        )
    # Drop bool col — not useful in Excel
    overview = overview.drop(columns=["G1_too_few"])
    return overview

# ── SHARED EXCEL HELPERS ──────────────────────────────────────────────────────
def write_title_row(ws, row, text, ncols):
    lc = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font      = bf(C_WHITE, 12)
    ws[f"A{row}"].fill      = xf(C_DARK)
    ws[f"A{row}"].alignment = al(1)
    ws.row_dimensions[row].height = 22

def write_sub_row(ws, row, text, ncols):
    lc = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font      = nf("444444", 8)
    ws[f"A{row}"].fill      = xf(C_LIGHT)
    ws[f"A{row}"].alignment = al(1)
    ws.row_dimensions[row].height = 13

def write_section_row(ws, row, text, ncols):
    lc = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{lc}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font      = bf(C_WHITE, 10)
    ws[f"A{row}"].fill      = xf(C_MID)
    ws[f"A{row}"].alignment = al(1)
    ws.row_dimensions[row].height = 17

def write_header_row(ws, row, headers, group_col_count=1):
    """Write header row with colour-coded group columns."""
    G_COLS = ["2E75B6","1F6B38","7F3F98","B85C00","B00020","1F3864"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        col_name = str(h)
        # Pick fill by prefix
        fill_c = C_MID
        for gi, prefix in enumerate(["G1","G2","G3","G4","G5","G6"]):
            if col_name.startswith(prefix): fill_c = G_COLS[gi]; break
        if col_name == "WHY_G1_vs_G2": fill_c = "555555"
        cell.font      = bf(C_WHITE, 9)
        cell.fill      = xf(fill_c)
        cell.border    = xb()
        cell.alignment = awrap()
    ws.row_dimensions[row].height = 36

def apply_cond_formats(ws, data_start_row, data_end_row, col_map: dict):
    """
    Apply conditional formatting rules to whole column ranges at once.
    col_map: {col_letter: rule_type}
    rule_type: 'pct_outlier' | 'pct_too_few' | 'outlier_count' | 'sufficient'
    """
    # Red fill for high outlier % (>=10)
    red_dxf    = DifferentialStyle(fill=xf(C_RED),
                                   font=Font(bold=True, color=C_RED_D, name="Arial", size=9))
    orange_dxf = DifferentialStyle(fill=xf(C_ORANGE),
                                   font=Font(bold=True, color="7F3F00", name="Arial", size=9))
    amber_dxf  = DifferentialStyle(fill=xf(C_AMBER),
                                   font=Font(bold=True, color="7F6000", name="Arial", size=9))
    green_dxf  = DifferentialStyle(fill=xf(C_GREEN),
                                   font=Font(bold=True, color=C_GREEN_D, name="Arial", size=9))

    for col_letter, rule_type in col_map.items():
        cell_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        if rule_type == "pct_outlier":
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="greaterThanOrEqual",
                     formula=["10"], dxf=red_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="between",
                     formula=["5","9.99"], dxf=orange_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="between",
                     formula=["2","4.99"], dxf=amber_dxf))
        elif rule_type == "pct_too_few":
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="greaterThan",
                     formula=["80"], dxf=red_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="between",
                     formula=["50","80"], dxf=amber_dxf))
        elif rule_type == "outlier_count":
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="greaterThan",
                     formula=["0"], dxf=red_dxf))
        elif rule_type == "sufficient_no":
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal",
                     formula=['"TOO FEW"'], dxf=amber_dxf))
        elif rule_type == "sufficient_yes":
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal",
                     formula=['"ENOUGH"'], dxf=green_dxf))

# ── STEP 5: OVERVIEW SHEET ────────────────────────────────────────────────────
def write_overview(wb, overview_df: pd.DataFrame):
    ws = wb.create_sheet("01_Overview")
    ncols = len(overview_df.columns)
    n_data = len(overview_df)

    write_title_row(ws, 1, "OUTLIER COMPARISON — All 6 grouping levels side by side, one row per commodity", ncols)
    write_sub_row(ws, 2,
        f"G1=Commodity | G2=+Country | G3=+Province | G4=+Country+Province | "
        f"G5=+Season | G6=+Country+Province+Season  |  TOO FEW=<{MIN_ROWS} valid rows  |  "
        f"Log-IQR ×{LOG_IQR_MULT} fences (IQR on log10 price — lower fence always positive)  |  Red=≥10% outlier | Orange=≥5% | Amber=≥2%",
        ncols)

    # Header row (row 3)
    write_header_row(ws, 3, list(overview_df.columns))

    # Data rows — bulk append, NO per-cell formatting
    for row_vals in dataframe_to_rows(overview_df, index=False, header=False):
        ws.append(row_vals)

    # Freeze + autofilter
    ws.freeze_panes = ws.cell(row=4, column=2)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"

    # Conditional formatting — one rule per column type, applied to whole range
    data_start = 4
    data_end   = 3 + n_data
    cond_map = {}
    for ci, col in enumerate(overview_df.columns, 1):
        ltr = get_column_letter(ci)
        if "pct_outlier" in col:  cond_map[ltr] = "pct_outlier"
        elif "pct_too_few" in col: cond_map[ltr] = "pct_too_few"
        elif col.endswith("_outliers") and col != "G1_outliers":
            cond_map[ltr] = "outlier_count"
    apply_cond_formats(ws, data_start, data_end, cond_map)

    # Column widths
    for ci, col in enumerate(overview_df.columns, 1):
        w = 52 if col == "Commodity" else \
            36 if col == "WHY_G1_vs_G2" else \
            13 if "valid_rows" in col else \
            12 if "groups" in col else \
            13 if "too_few" in col else \
            13 if "outlier" in col else 12
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Row heights — uniform for data rows
    for r in range(4, data_end + 1):
        ws.row_dimensions[r].height = 15

    print(f"  Sheet '01_Overview': {n_data:,} rows")

# ── STEP 6: DETAIL SHEET — FAST ──────────────────────────────────────────────
def write_detail_sheet(wb, sheet_name: str, df_g: pd.DataFrame,
                       group_cols: list, title: str):
    # Rename columns for display
    rename = {
        "n_rows":       "Valid Rows",
        "total_value":  "Total Value ($)",
        "total_qty":    "Total Qty (normalised)",
        "sufficient":   "Sufficient",
        "median_price": "Median Unit Price", "mean_price": "Mean Unit Price",
        "std_price": "Std Dev", "cv": "CV",
        "min_price": "Min Unit Price", "max_price": "Max Unit Price",
        "q1": "Q1", "q3": "Q3", "iqr": "IQR",
        "lower_fence": f"Lower Fence (log-IQR ×{LOG_IQR_MULT})",
        "upper_fence": f"Upper Fence (log-IQR ×{LOG_IQR_MULT})",
        "n_outlier_any": "Outliers",
        "pct_outlier":   "% Outlier",
        "n_outlier_med": "MEDIUM",
        "n_outlier_hi":  "HIGH",
        "n_outlier_crit":"CRITICAL",
        "n_heterogeneous": "REVIEW (>15% flagged)",
        "pct_no_outl":   "% Clean",
        "too_few":       "Status",
        "_season":       "Season",
    }
    display_cols = (group_cols +
        ["n_rows", "total_value", "total_qty", "sufficient",
         "median_price","mean_price","std_price","cv",
         "min_price","max_price","q1","q3","iqr",
         "lower_fence","upper_fence",
         "n_outlier_any","pct_outlier",
         "n_outlier_med","n_outlier_hi","n_outlier_crit",
         "n_heterogeneous",
         "pct_no_outl"])
    display_cols = [c for c in display_cols if c in df_g.columns]

    df_out = df_g[display_cols].copy()
    # Convert bool sufficient → readable string
    df_out["sufficient"] = df_out["sufficient"].map({True: "ENOUGH", False: "TOO FEW"})
    df_out = df_out.rename(columns=rename)

    ws = wb.create_sheet(sheet_name)
    ncols  = len(df_out.columns)
    n_data = len(df_out)

    write_title_row(ws, 1, title, ncols)
    write_sub_row(ws, 2,
        f"Log-IQR ×{LOG_IQR_MULT}: fences on log10(price), always positive | "
        f"Amber = TOO FEW (<{MIN_ROWS} valid rows) | Green = ENOUGH | "
        f"Red % Outlier ≥10% | Orange ≥5%", ncols)

    write_header_row(ws, 3, list(df_out.columns))

    # Bulk append all data rows
    for row_vals in dataframe_to_rows(df_out, index=False, header=False):
        ws.append(row_vals)

    ws.freeze_panes = ws.cell(row=4, column=len(group_cols) + 1)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"

    # Conditional formatting rules — applied once to full column range
    data_end = 3 + n_data
    cond_map = {}
    for ci, col in enumerate(df_out.columns, 1):
        ltr = get_column_letter(ci)
        if col == "% Outlier":      cond_map[ltr] = "pct_outlier"
        elif col == "Status":
            cond_map[ltr] = "sufficient_no"
            apply_cond_formats(ws, 4, data_end, {ltr: "sufficient_yes"})
        elif col in ("Outliers","MEDIUM","HIGH","CRITICAL","REVIEW (>15% flagged)"):
            cond_map[ltr] = "outlier_count"
    apply_cond_formats(ws, 4, data_end, cond_map)

    # Column widths
    widths = []
    for col in df_out.columns:
        if   col == "Commodity":                    widths.append(50)
        elif col == "Country":                      widths.append(24)
        elif col in ("Province","Season"):          widths.append(18)
        elif col == "Status":                       widths.append(11)
        elif col in ("Valid Rows","CV"):            widths.append(11)
        elif col == "Total Value ($)":              widths.append(18)
        elif col == "Total Qty (normalised)":       widths.append(18)
        elif "Unit Price" in col:                   widths.append(16)
        elif "Fence" in col or col[:3] == "IQR":   widths.append(15)
        elif "%" in col:                            widths.append(12)
        elif any(x in col for x in ("Outlier","MEDIUM","HIGH","CRITICAL","REVIEW")): widths.append(14)
        else:                             widths.append(13)
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Uniform row height — much faster than setting each row
    ws.sheet_format.defaultRowHeight = 15
    ws.sheet_format.customHeight     = True

    print(f"  Sheet '{sheet_name}': {n_data:,} groups")

# ── STEP 7: RAW COUNTS SHEET ──────────────────────────────────────────────────
def write_raw_counts(wb, all_stats: dict):
    ws = wb.create_sheet("08_Raw_Counts")
    ncols = 10
    write_title_row(ws, 1,
        "GROUP SIZE DISTRIBUTION — how data thins as grouping becomes more granular", ncols)
    write_sub_row(ws, 2,
        "Median group size of 1-2 rows = most groups are statistically useless. "
        "Red % TOO FEW > 80% = grouping level is too granular for outlier detection.", ncols)

    headers = ["Grouping Level","Group Columns","Total Groups",
               f"TOO FEW (<{MIN_ROWS})","% TOO FEW",
               "Median group size","Groups ≥5","Groups ≥10","Groups ≥20","Groups ≥50"]
    write_header_row(ws, 3, headers)

    label_nice = {
        "G1_Commodity":              "G1 — Commodity only",
        "G2_Comm_Country":           "G2 — Commodity + Country",
        "G3_Comm_Province":          "G3 — Commodity + Province",
        "G4_Comm_Country_Province":  "G4 — Commodity + Country + Province",
        "G5_Comm_Season":            "G5 — Commodity + Season",
        "G6_Comm_Country_Prov_Seas": "G6 — Commodity + Country + Province + Season",
    }
    rows = []
    for gname, df_g in all_stats.items():
        n        = len(df_g)
        too_few  = int((~df_g["sufficient"]).sum())
        pct_tf   = round(too_few / n * 100, 1) if n else 0
        med_sz   = round(float(df_g["n_rows"].median()), 1)
        gcols    = " + ".join(GROUPINGS[gname])
        rows.append([
            label_nice[gname], gcols, n, too_few, pct_tf,
            med_sz,
            int((df_g["n_rows"] >= 5).sum()),
            int((df_g["n_rows"] >= 10).sum()),
            int((df_g["n_rows"] >= 20).sum()),
            int((df_g["n_rows"] >= 50).sum()),
        ])
    for r in rows:
        ws.append(r)

    # Cond format on % TOO FEW column (col E = 5)
    apply_cond_formats(ws, 4, 3 + len(rows), {"E": "pct_too_few"})
    set_col_widths = [30, 48, 13, 13, 12, 18, 12, 12, 12, 12]
    for ci, w in enumerate(set_col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.sheet_format.defaultRowHeight = 15

    print(f"  Sheet '08_Raw_Counts': {len(rows)} grouping levels")

# ── STEP 8: HOW TO READ ───────────────────────────────────────────────────────
def write_how_to(wb):
    ws = wb.create_sheet("00_How_To_Read", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 68

    rows = [
        ("HOW TO READ THIS FILE", None, "title"),
        ("PURPOSE", None, "sec"),
        ("This file answers one question: at what grouping level should outliers be detected?",
         "It shows outlier counts across all 6 grouping levels BEFORE committing to any strategy.", "txt"),
        ("", "", "blank"),
        ("SHEETS", None, "sec"),
        ("01_Overview",         "One row per commodity. Outlier counts at every level side by side. START HERE.", "col"),
        ("02_G1_Commodity",     "Full stats — Commodity only (baseline)", "col"),
        ("03_G2_CommCountry",   "Full stats — Commodity + Country", "col"),
        ("04_G3_CommProvince",  "Full stats — Commodity + Province", "col"),
        ("05_G4_CommCntryProv", "Full stats — Commodity + Country + Province", "col"),
        ("06_G5_CommSeason",    "Full stats — Commodity + Season (Q1-Q4)", "col"),
        ("07_G6_AllGroups",     "Full stats — Commodity + Country + Province + Season", "col"),
        ("08_Raw_Counts",       "Group size distribution — shows exactly how data thins at each level", "col"),
        ("", "", "blank"),
        ("HOW TO USE 01_Overview", None, "sec"),
        ("Step 1 — Baseline",   "G1_pct_outlier = your baseline outlier rate at commodity level.", "col"),
        ("Step 2 — Compare",    "Look at G2_pct_outlier vs G1. Did it drop? Now check G2_pct_too_few. If TOO FEW went way up, the drop is fragmentation — not a real improvement.", "col"),
        ("Step 3 — WHY column", "'Better precision' = real improvement (group has enough rows and stats are tighter). 'Fragmentation' = artificial drop because groups are too small to detect anything.", "col"),
        ("Step 4 — 08_Raw_Counts", "Check median group size at each level. Median of 1-2 rows = statistically useless.", "col"),
        ("", "", "blank"),
        ("COLOUR GUIDE", None, "sec"),
        ("% Outlier — Red",    "≥10% of rows flagged. Likely a systematic issue or natural price variation.", "col"),
        ("% Outlier — Orange", "5-9.9% flagged. Worth investigating.", "col"),
        ("% Outlier — Amber",  "2-4.9% flagged. Moderate.", "col"),
        ("% TOO FEW — Red",    ">80% of groups have insufficient data. This grouping level is too granular.", "col"),
        ("% TOO FEW — Amber",  "50-80%. Use with caution.", "col"),
        ("Status — Amber",     "TOO FEW: fewer than 5 valid price rows. Statistics unreliable.", "col"),
        ("Status — Green",     "ENOUGH: 5+ valid rows. Statistics are reliable.", "col"),
        ("", "", "blank"),
        ("INTERPRETING OUTLIER CHANGES BETWEEN LEVELS", None, "sec"),
        ("Outliers DOWN + TOO FEW UP",   "Fragmentation. Stats can't fire because groups too small. Not a real improvement.", "col"),
        ("Outliers DOWN + TOO FEW low",  "Better precision. Subgroup is large enough, median/std became more representative.", "col"),
        ("Outliers UP after grouping",   "Unusual. A subgroup may have extreme prices masked in the global distribution.", "col"),
        ("Outliers unchanged",           "Grouping doesn't help for this commodity. Use the simpler group.", "col"),
    ]

    for i, (label, value, style) in enumerate(rows, 1):
        ws.row_dimensions[i].height = 16
        if style == "title":
            ws.merge_cells(f"A{i}:B{i}")
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = bf(C_WHITE, 13); ws[f"A{i}"].fill = xf(C_DARK)
            ws[f"A{i}"].alignment = al(1); ws.row_dimensions[i].height = 26
        elif style == "sec":
            ws.merge_cells(f"A{i}:B{i}")
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = bf(C_WHITE, 10); ws[f"A{i}"].fill = xf(C_MID)
            ws[f"A{i}"].alignment = al(1); ws.row_dimensions[i].height = 18
        elif style == "col":
            ws[f"A{i}"] = label; ws[f"B{i}"] = value
            ws[f"A{i}"].font = bf("000000", 9); ws[f"B{i}"].font = nf("000000", 9)
            ws[f"A{i}"].fill = xf(C_GREY);      ws[f"B{i}"].fill = xf(C_GREY)
            ws[f"A{i}"].alignment = al(1);       ws[f"B{i}"].alignment = al(1)
        elif style == "txt":
            ws[f"A{i}"] = label; ws[f"B{i}"] = value
            ws[f"A{i}"].font = nf("000000", 9); ws[f"B{i}"].font = nf("000000", 9)
            ws[f"A{i}"].alignment = al(2);      ws[f"B{i}"].alignment = al(1)
        elif style == "blank":
            ws.merge_cells(f"A{i}:B{i}")

# ── MAIN ──────────────────────────────────────────────────────────────────────

# ── STEP: BUILD FLAGGED ROWS EXPORT ──────────────────────────────────────────
def build_flagged_rows(df_valid: pd.DataFrame, all_flagged: dict) -> pd.DataFrame:
    """
    Memory-safe version: each grouping contributes only a slim
    [_row_id, severity, z, median, fences] dataframe (~3-5 MB each).
    We join on _row_id (integer index) — no large merges on string columns.

    FLAG_COUNT = how many of the 6 grouping levels flagged each row.
    """
    print("  Building flagged rows export...", end=" ", flush=True)
    t0 = time.time()

    short = {
        "G1_Commodity":              "G1",
        "G2_Comm_Country":           "G2",
        "G3_Comm_Province":          "G3",
        "G4_Comm_Country_Province":  "G4",
        "G5_Comm_Season":            "G5",
        "G6_Comm_Country_Prov_Seas": "G6",
    }

    # Build a combined flags table: one row per (_row_id, grouping)
    # Each all_flagged[gname] is already slim: [_row_id, _severity, _z, median_price, fences]
    flag_frames = []
    for gname, slim in all_flagged.items():
        p = short[gname]
        f = slim.rename(columns={
            "_severity":    f"{p}_severity",
            "_z":           f"{p}_z",
            "log_med_price": f"{p}_median_price",
            "lower_fence":   f"{p}_lower_fence",
            "upper_fence":   f"{p}_upper_fence",
        }).set_index("_row_id")
        flag_frames.append(f)
        del slim  # free immediately

    # Outer join all flag frames on _row_id — all are small, safe in memory
    flags_combined = flag_frames[0]
    for f in flag_frames[1:]:
        flags_combined = flags_combined.join(f, how="outer")
    del flag_frames

    # FLAG_COUNT
    sev_cols = [f"{p}_severity" for p in short.values() if f"{p}_severity" in flags_combined.columns]
    flags_combined["FLAG_COUNT"] = flags_combined[sev_cols].notna().sum(axis=1).astype(int)

    # Now join back to original data — only for flagged rows (inner join = keeps only matches)
    base_cols = ["_row_id", "Commodity", "Period", "Country", "Province", "_season",
                 "Value ($)", "Quantity", "Unit of measure", "_base_unit",
                 "_unit_price", "_source_file"]
    base_cols = [col for col in base_cols if col in df_valid.columns]
    base = df_valid[base_cols].set_index("_row_id")

    result = base.join(flags_combined, how="inner").reset_index(drop=True)
    del base, flags_combined

    # Clean display order
    ordered = (
        ["Commodity", "Period", "Country", "Province", "_season",
         "Value ($)", "Quantity", "Unit of measure", "_base_unit",
         "_unit_price", "_source_file", "FLAG_COUNT"] +
        [f"{p}_{s}" for p in short.values()
                    for s in ["severity", "z", "median_price", "lower_fence", "upper_fence"]]
    )
    ordered = [col for col in ordered if col in result.columns]
    result = (result[ordered]
              .rename(columns={"_season": "Season", "_unit_price": "Unit Price",
                               "_base_unit": "Base Unit", "_source_file": "Source File"})
              .sort_values(["Commodity", "Country"], na_position="last")
              .reset_index(drop=True))

    print(f"{len(result):,} flagged rows  ({time.time()-t0:.1f}s)")
    return result


def write_flagged_rows_sheet(wb, flagged_df: pd.DataFrame):
    """Fast sheet: flagged rows with conditional formatting on severity."""
    ws = wb.create_sheet("09_Flagged_Rows")
    ncols  = len(flagged_df.columns)
    n_data = len(flagged_df)

    write_title_row(ws, 1,
        f"FLAGGED ROWS — {n_data:,} rows flagged by at least one grouping level", ncols)
    write_sub_row(ws, 2,
        "FLAG_COUNT = how many of the 6 grouping levels flagged this row  |  "
        "G1=Commodity | G2=+Country | G3=+Province | G4=+Country+Province | "
        "G5=+Season | G6=+Country+Province+Season  |  "
        "Severity (v4 guards): CRITICAL/HIGH/MEDIUM = percentile-rank within group outliers  |  "
        "REVIEW = group outlier rate >15% (structurally heterogeneous)  |  "
        "CV<2% groups fully suppressed  |  Min log-std floor prevents z-score inflation", ncols)

    write_header_row(ws, 3, list(flagged_df.columns))

    # Bulk append rows
    for row_vals in dataframe_to_rows(flagged_df, index=False, header=False):
        # Convert bool to YES/NO for readability
        clean = ["YES" if v is True else ("NO" if v is False else v) for v in row_vals]
        ws.append(clean)

    ws.freeze_panes = ws.cell(row=4, column=2)
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"
    ws.sheet_format.defaultRowHeight = 15

    # Conditional formatting
    data_end = 3 + n_data
    cond_map = {}
    red_dxf    = DifferentialStyle(fill=xf(C_RED),
                                   font=Font(bold=True, color=C_RED_D, name="Arial", size=9))
    amber_dxf  = DifferentialStyle(fill=xf(C_AMBER),
                                   font=Font(bold=True, color="7F6000", name="Arial", size=9))
    orange_dxf = DifferentialStyle(fill=xf(C_ORANGE),
                                   font=Font(bold=True, color="7F3F00", name="Arial", size=9))
    green_dxf  = DifferentialStyle(fill=xf(C_GREEN),
                                   font=Font(bold=True, color=C_GREEN_D, name="Arial", size=9))

    for ci, col in enumerate(flagged_df.columns, 1):
        ltr = get_column_letter(ci)
        cell_range = f"{ltr}4:{ltr}{data_end}"
        if col == "FLAG_COUNT":
            # Colour by count: 1=amber, 2-3=orange, 4+=red
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="greaterThanOrEqual", formula=["4"], dxf=red_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="between", formula=["2","3"], dxf=orange_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=["1"], dxf=amber_dxf))
        elif col.endswith("_severity"):
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=['"CRITICAL"'], dxf=red_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=['"HIGH"'], dxf=orange_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=['"MEDIUM"'], dxf=amber_dxf))
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=['"REVIEW"'], dxf=green_dxf))
        elif col.endswith("_flagged"):
            ws.conditional_formatting.add(cell_range,
                Rule(type="cellIs", operator="equal", formula=['"YES"'], dxf=red_dxf))

    # Column widths
    for ci, col in enumerate(flagged_df.columns, 1):
        w = (52 if col == "Commodity" else
             10 if col == "Period" else
             24 if col == "Country" else
             18 if col in ("Province","Season") else
             10 if col == "FLAG_COUNT" else
             26 if col == "Source File" else
             10 if col in ("Value ($)","Quantity","Unit Price") else
             10 if col in ("Base Unit","Unit of measure") else
             9  if col.endswith("_flagged") else
             12 if col.endswith("_severity") else
             9  if col.endswith("_z") else
             13 if "price" in col.lower() or "fence" in col.lower() else 11)
        ws.column_dimensions[get_column_letter(ci)].width = w

    print(f"  Sheet '09_Flagged_Rows': {n_data:,} rows")


def export_flagged_csv(flagged_df: pd.DataFrame, output_dir: Path):
    """Also save as CSV — much faster to open in Excel than a large .xlsx sheet."""
    csv_path = output_dir / "flagged_rows.csv"
    flagged_df.to_csv(csv_path, index=False)
    mb = csv_path.stat().st_size / 1_048_576
    print(f"  CSV export: {csv_path.name}  ({mb:.1f} MB)")
    return csv_path

def main(input_path: Path):
    S = "=" * 65
    print(f"\n{S}\n  OUTLIER GROUPING DIAGNOSTIC v3 (log-IQR)\n{S}\n")

    df_all, df_valid = load(input_path)

    print(f"\n  Computing stats — {len(GROUPINGS)} grouping levels...")
    all_stats    = {}
    all_flagged  = {}
    total_outliers_summary = []
    for gname, gcols in GROUPINGS.items():
        stats, flagged_rows, elapsed = compute_group_stats(df_valid, gcols)
        n_too_few = int((~stats["sufficient"]).sum())
        pct_tf    = n_too_few / len(stats) * 100
        n_out     = int(stats["n_outlier_any"].sum())
        print(f"  {gname:<30} {len(stats):>7,} groups | "
              f"TOO FEW: {n_too_few:>6,} ({pct_tf:5.1f}%) | "
              f"Outliers: {n_out:>6,} | {elapsed:.1f}s")
        all_stats[gname]   = stats
        all_flagged[gname] = flagged_rows
        total_outliers_summary.append((gname, len(stats), n_too_few, pct_tf, n_out))

    print("\n  Summary — outlier count by grouping level:")
    print(f"  {'Grouping':<32} {'Groups':>8} {'TOO FEW%':>10} {'Outliers':>10}")
    print("  " + "-"*62)
    for gname, ng, ntf, ptf, no in total_outliers_summary:
        print(f"  {gname:<32} {ng:>8,} {ptf:>9.1f}% {no:>10,}")

    print("\n  Building overview...")
    overview = build_overview(all_stats)

    print("\n  Writing workbook (fast mode — conditional formatting)...")
    t_write = time.time()
    wb = Workbook()
    wb.remove(wb.active)

    write_how_to(wb)
    write_overview(wb, overview)

    sheet_map = {
        "G1_Commodity":             ("02_G1_Commodity",     f"G1: Commodity only — {len(all_stats['G1_Commodity']):,} groups"),
        "G2_Comm_Country":          ("03_G2_CommCountry",   f"G2: Commodity + Country — {len(all_stats['G2_Comm_Country']):,} groups"),
        "G3_Comm_Province":         ("04_G3_CommProvince",  f"G3: Commodity + Province — {len(all_stats['G3_Comm_Province']):,} groups"),
        "G4_Comm_Country_Province": ("05_G4_CommCntryProv", f"G4: Commodity + Country + Province — {len(all_stats['G4_Comm_Country_Province']):,} groups"),
        "G5_Comm_Season":           ("06_G5_CommSeason",    f"G5: Commodity + Season — {len(all_stats['G5_Comm_Season']):,} groups"),
        "G6_Comm_Country_Prov_Seas":("07_G6_AllGroups",     f"G6: Commodity + Country + Province + Season — {len(all_stats['G6_Comm_Country_Prov_Seas']):,} groups"),
    }
    for gname, (sname, title) in sheet_map.items():
        write_detail_sheet(wb, sname, all_stats[gname], GROUPINGS[gname], title)

    write_raw_counts(wb, all_stats)

    # Flagged rows export
    flagged_df = build_flagged_rows(df_valid, all_flagged)
    del all_flagged  # free ~30MB of slim flag frames
    write_flagged_rows_sheet(wb, flagged_df)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print("  Saving file...", end=" ", flush=True)
    t_save = time.time()
    wb.save(OUTPUT_FILE)
    mb = OUTPUT_FILE.stat().st_size / 1_048_576
    print(f"done ({time.time()-t_save:.1f}s) — {mb:.1f} MB")
    print(f"  Excel write total: {time.time()-t_write:.1f}s")
    export_flagged_csv(flagged_df, OUTPUT_DIR)
    print(f"\n  Saved: {OUTPUT_FILE}")
    print(f"\n{S}\n  DONE\n{S}\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=INPUT_FILE)
    args = parser.parse_args()
    if not args.input.exists():
        print(f"\n  ERROR: File not found: {args.input}")
        raise SystemExit(1)
    main(args.input)
